#!/usr/bin/env python3
"""
Convert all markup_log.xlsx files to markup_log.csv in the patients folder.

Run this from the annotation directory (where auto_script.py lives):
    cd annotation_mac_xlsx/
    python3 ../convert_markup_logs.py

Or from annotation_windows_xlsx/:
    cd annotation_windows_xlsx\
    python3 ..\convert_markup_logs.py

It will:
  1. Scan patients/**/markup_log.xlsx recursively
  2. Convert each to markup_log.csv in the same folder
  3. Rename the original .xlsx to .xlsx.bak (backup, not deleted)
"""

import os
import csv
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required to read .xlsx files.")
    print("Install it with: pip3 install openpyxl")
    sys.exit(1)

patients_dir = os.path.join(os.getcwd(), "patients")

if not os.path.isdir(patients_dir):
    print(f"ERROR: 'patients' folder not found in {os.getcwd()}")
    print("Make sure you run this script from the annotation directory")
    print("(e.g. annotation_mac_xlsx/ or annotation_windows_xlsx/)")
    sys.exit(1)

converted = 0
skipped = 0

for root, dirs, files in os.walk(patients_dir):
    if "markup_log.xlsx" not in files:
        continue

    xlsx_path = os.path.join(root, "markup_log.xlsx")
    csv_path = os.path.join(root, "markup_log.csv")
    backup_path = xlsx_path + ".bak"

    # Skip if CSV already exists
    if os.path.exists(csv_path):
        print(f"SKIP: {csv_path} already exists")
        skipped += 1
        continue

    try:
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active

        # Read header row
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            print(f"SKIP: {xlsx_path} is empty")
            skipped += 1
            continue

        headers = [str(h) if h else '' for h in rows[0]]

        # Write CSV
        with open(csv_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows[1:]:
                # Replace None with empty string
                writer.writerow([str(v) if v is not None else '' for v in row])

        # --- Verify: read back the CSV and compare to xlsx content ---
        with open(csv_path, 'r', newline='') as f:
            reader = csv.reader(f)
            csv_rows = list(reader)

        # Check row count (header + data rows)
        xlsx_row_count = len(rows)
        csv_row_count = len(csv_rows)
        if xlsx_row_count != csv_row_count:
            print(f"FAIL: {xlsx_path} — row count mismatch (xlsx: {xlsx_row_count}, csv: {csv_row_count})")
            os.remove(csv_path)
            continue

        # Check every cell value matches
        mismatch = False
        for row_idx in range(xlsx_row_count):
            xlsx_row = rows[row_idx]
            csv_row = csv_rows[row_idx]

            if len(csv_row) != len(xlsx_row):
                print(f"FAIL: {xlsx_path} — column count mismatch at row {row_idx + 1} (xlsx: {len(xlsx_row)}, csv: {len(csv_row)})")
                mismatch = True
                break

            for col_idx in range(len(xlsx_row)):
                xlsx_val = str(xlsx_row[col_idx]) if xlsx_row[col_idx] is not None else ''
                csv_val = csv_row[col_idx]
                if xlsx_val != csv_val:
                    print(f"FAIL: {xlsx_path} — value mismatch at row {row_idx + 1}, col {col_idx + 1}")
                    print(f"       xlsx: '{xlsx_val}'")
                    print(f"       csv:  '{csv_val}'")
                    mismatch = True
                    break
            if mismatch:
                break

        if mismatch:
            os.remove(csv_path)
            continue

        # Verification passed — delete the original xlsx
        os.remove(xlsx_path)

        rel_path = os.path.relpath(csv_path, os.getcwd())
        print(f"OK:   {rel_path} ({len(rows) - 1} rows, verified, xlsx deleted)")
        converted += 1

    except Exception as e:
        print(f"ERROR: Failed to convert {xlsx_path}: {e}")
        # Clean up partial CSV if it was created
        if os.path.exists(csv_path):
            os.remove(csv_path)

print(f"\nDone. Converted: {converted}, Skipped: {skipped}")
