#!/usr/bin/env python3
"""
Unit tests for annotation workflow integrity.

Tests verify that:
1. log.csv contains all patient scanning sessions
2. All paths in log.csv point to valid directories
3. log.csv format is correct
4. Patient data structure is consistent

Usage:
    # Test workflows in current directory
    python test_annotation_workflow.py

    # Test specific directory containing workflows
    python test_annotation_workflow.py --directory /path/to/workflows

    # Test a single workflow directory
    python test_annotation_workflow.py --workflow /path/to/annotation_mac_xlsx

    # Test with verbose output
    python test_annotation_workflow.py -v
"""

import unittest
import csv
import re
import argparse
import sys
import json
from pathlib import Path
from datetime import datetime

# Global variable to store custom directory path
CUSTOM_DIRECTORY = None
CUSTOM_WORKFLOW = None

# Global variable to store deleted markups information
DELETED_MARKUPS_LOG = []


class TestAnnotationWorkflowIntegrity(unittest.TestCase):
    """Test suite for annotation workflow data integrity."""

    def setUp(self):
        """Set up test fixtures - find workflow directories."""
        global CUSTOM_DIRECTORY, CUSTOM_WORKFLOW

        self.workflows = []

        # If testing a single workflow directory
        if CUSTOM_WORKFLOW:
            workflow_path = Path(CUSTOM_WORKFLOW)
            if workflow_path.exists() and workflow_path.is_dir():
                # Check if this looks like a workflow directory (has log.csv and patients)
                if (workflow_path / 'log.csv').exists() or (workflow_path / 'patients').exists():
                    self.workflows.append({
                        'name': workflow_path.name,
                        'path': workflow_path,
                        'log_csv': workflow_path / 'log.csv',
                        'patients_dir': workflow_path / 'patients'
                    })
            return

        # Determine the base directory to search
        if CUSTOM_DIRECTORY:
            base_dir = Path(CUSTOM_DIRECTORY)
        else:
            base_dir = Path(__file__).parent

        # Find all annotation workflow directories
        # First, try known workflow names
        for workflow_dir in ['annotation_mac_xlsx', 'annotation_windows_xlsx',
                             'annotation_mac', 'annotation_windows']:
            workflow_path = base_dir / workflow_dir
            if workflow_path.exists() and workflow_path.is_dir():
                self.workflows.append({
                    'name': workflow_dir,
                    'path': workflow_path,
                    'log_csv': workflow_path / 'log.csv',
                    'patients_dir': workflow_path / 'patients'
                })

        # If no known workflows found, search for any directory with log.csv and patients
        if not self.workflows:
            for item in base_dir.iterdir():
                if item.is_dir():
                    log_csv = item / 'log.csv'
                    patients_dir = item / 'patients'
                    if log_csv.exists() or patients_dir.exists():
                        self.workflows.append({
                            'name': item.name,
                            'path': item,
                            'log_csv': log_csv,
                            'patients_dir': patients_dir
                        })

    def test_workflows_exist(self):
        """Test that at least one workflow directory exists."""
        self.assertGreater(len(self.workflows), 0,
                          "No annotation workflow directories found")

    def test_log_csv_exists(self):
        """Test that log.csv exists in each workflow."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                self.assertTrue(
                    workflow['log_csv'].exists(),
                    f"log.csv not found in {workflow['name']}"
                )

    def test_patients_directory_exists(self):
        """Test that patients directory exists in each workflow."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                self.assertTrue(
                    workflow['patients_dir'].exists(),
                    f"patients directory not found in {workflow['name']}"
                )

    def test_log_csv_format(self):
        """Test that log.csv has the correct format (Report,Path,Done header)."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.reader(f)
                    header = next(reader)

                    self.assertEqual(len(header), 3,
                                   f"log.csv should have 3 columns in {workflow['name']}")
                    self.assertEqual(header[0], 'Report',
                                   f"First column should be 'Report' in {workflow['name']}")
                    self.assertEqual(header[1], 'Path',
                                   f"Second column should be 'Path' in {workflow['name']}")
                    self.assertEqual(header[2], 'Done',
                                   f"Third column should be 'Done' in {workflow['name']}")

    def test_log_csv_entries_valid(self):
        """Test that all entries in log.csv have valid report numbers and paths."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row_num, row in enumerate(reader, start=2):  # Start at 2 (after header)
                        # Test report number is valid
                        report = row.get('Report', '').strip()
                        self.assertTrue(
                            report.isdigit(),
                            f"Row {row_num} in {workflow['name']}: Report '{report}' is not a valid number"
                        )

                        # Test path is not empty
                        path = row.get('Path', '').strip()
                        self.assertNotEqual(
                            path, '',
                            f"Row {row_num} in {workflow['name']}: Path is empty"
                        )

    def test_log_csv_paths_exist(self):
        """Test that all paths in log.csv point to existing directories."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        report_num = row.get('Report', '').strip()
                        relative_path = row.get('Path', '').strip()

                        # Convert relative path to absolute
                        # Handle both ./ and direct paths
                        if relative_path.startswith('./'):
                            relative_path = relative_path[2:]

                        abs_path = workflow['path'] / relative_path

                        self.assertTrue(
                            abs_path.exists(),
                            f"Report {report_num} in {workflow['name']}: Path '{relative_path}' does not exist"
                        )
                        self.assertTrue(
                            abs_path.is_dir(),
                            f"Report {report_num} in {workflow['name']}: Path '{relative_path}' is not a directory"
                        )

    def test_all_scanning_sessions_in_log(self):
        """Test that all scanning_session directories are listed in log.csv."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['patients_dir'].exists():
                    self.skipTest(f"patients directory not found in {workflow['name']}")
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                # Find all scanning_session directories
                scanning_sessions = []
                for patient_dir in workflow['patients_dir'].iterdir():
                    if patient_dir.is_dir() and patient_dir.name.startswith('patient'):
                        for session_dir in patient_dir.iterdir():
                            if session_dir.is_dir() and session_dir.name.startswith('scanning_session'):
                                # Get relative path from workflow root
                                rel_path = session_dir.relative_to(workflow['path'])
                                scanning_sessions.append(str(rel_path))

                # Read all paths from log.csv
                log_paths = []
                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        path = row.get('Path', '').strip()
                        # Normalize path (remove ./ prefix if present)
                        if path.startswith('./'):
                            path = path[2:]
                        log_paths.append(path)

                # Check that all scanning sessions are in log
                for session in scanning_sessions:
                    self.assertIn(
                        session,
                        log_paths,
                        f"Scanning session '{session}' not found in log.csv of {workflow['name']}"
                    )

    def test_no_duplicate_reports(self):
        """Test that there are no duplicate report numbers in log.csv."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                report_numbers = []
                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        report = row.get('Report', '').strip()
                        report_numbers.append(report)

                # Check for duplicates
                duplicates = [num for num in report_numbers if report_numbers.count(num) > 1]
                unique_duplicates = list(set(duplicates))

                self.assertEqual(
                    len(unique_duplicates), 0,
                    f"Duplicate report numbers found in {workflow['name']}: {unique_duplicates}"
                )

    def test_scanning_sessions_have_required_files(self):
        """Test that each scanning session has at least one NIfTI file and one report file."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        report_num = row.get('Report', '').strip()
                        relative_path = row.get('Path', '').strip()

                        # Convert relative path to absolute
                        if relative_path.startswith('./'):
                            relative_path = relative_path[2:]

                        session_dir = workflow['path'] / relative_path

                        if not session_dir.exists():
                            continue  # Skip if directory doesn't exist (covered by other test)

                        # Check for NIfTI files
                        nifti_files = list(session_dir.glob('*.nii')) + list(session_dir.glob('*.nii.gz'))
                        self.assertGreater(
                            len(nifti_files), 0,
                            f"Report {report_num} in {workflow['name']}: No NIfTI files found in {relative_path}"
                        )

                        # Check for report text files
                        txt_files = list(session_dir.glob('report_*.txt'))
                        self.assertGreater(
                            len(txt_files), 0,
                            f"Report {report_num} in {workflow['name']}: No report text file found in {relative_path}"
                        )

    def test_report_numbers_are_sequential(self):
        """Test that report numbers in log.csv are sequential starting from 1."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                report_numbers = []
                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        report = row.get('Report', '').strip()
                        if report.isdigit():
                            report_numbers.append(int(report))

                if not report_numbers:
                    self.skipTest(f"No valid report numbers in {workflow['name']}")

                report_numbers.sort()

                # Check that numbers start at 1
                self.assertEqual(
                    report_numbers[0], 1,
                    f"Report numbers should start at 1 in {workflow['name']}"
                )

                # Check that numbers are sequential
                expected = list(range(1, len(report_numbers) + 1))
                self.assertEqual(
                    report_numbers, expected,
                    f"Report numbers are not sequential in {workflow['name']}: {report_numbers}"
                )

    def test_actual_report_files_exist(self):
        """Test that each scanning session has exactly one report_*.txt file with actual report number."""
        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        report_num = row.get('Report', '').strip()
                        relative_path = row.get('Path', '').strip()

                        # Convert relative path to absolute
                        if relative_path.startswith('./'):
                            relative_path = relative_path[2:]

                        session_dir = workflow['path'] / relative_path

                        if not session_dir.exists():
                            continue  # Skip if directory doesn't exist

                        # Find all report_*.txt files
                        report_files = list(session_dir.glob('report_*.txt'))

                        # Should have exactly one report file
                        self.assertEqual(
                            len(report_files), 1,
                            f"Report {report_num} in {workflow['name']}: Expected exactly 1 report_*.txt file in {relative_path}, found {len(report_files)}"
                        )

                        # Extract actual report number from filename
                        if report_files:
                            report_filename = report_files[0].name
                            # Extract number from report_13.txt -> 13
                            match = re.search(r'report_(\d+)\.txt', report_filename)
                            if match:
                                actual_report_num = match.group(1)
                                # Just verify it's a number (can be any value like 13, 418, etc.)
                                self.assertTrue(
                                    actual_report_num.isdigit(),
                                    f"Report {report_num} in {workflow['name']}: Actual report number '{actual_report_num}' in filename should be numeric"
                                )

    def test_log_deleted_markups(self):
        """Log all reports that have deleted markups in their markup_log files."""
        global DELETED_MARKUPS_LOG

        for workflow in self.workflows:
            with self.subTest(workflow=workflow['name']):
                if not workflow['log_csv'].exists():
                    self.skipTest(f"log.csv not found in {workflow['name']}")

                with open(workflow['log_csv'], 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        report_num = row.get('Report', '').strip()
                        relative_path = row.get('Path', '').strip()

                        # Convert relative path to absolute
                        if relative_path.startswith('./'):
                            relative_path = relative_path[2:]

                        session_dir = workflow['path'] / relative_path

                        if not session_dir.exists():
                            continue

                        # Check for markup_log.xlsx (XLSX version)
                        markup_log_xlsx = session_dir / 'markup_log.xlsx'
                        markup_log_csv = session_dir / 'markup_log.csv'

                        deleted_markups = []

                        # Try XLSX first (requires openpyxl)
                        if markup_log_xlsx.exists():
                            try:
                                import openpyxl
                                wb = openpyxl.load_workbook(markup_log_xlsx, read_only=True)
                                ws = wb.active

                                # Get header → column index mapping
                                headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=0)}

                                if 'deleted_at' in headers:
                                    deleted_at_col = headers['deleted_at']
                                    filename_col = headers.get('new_filename', 0)
                                    report_content_col = headers.get('report_content', 1)

                                    # Go through rows after header
                                    for row in ws.iter_rows(min_row=2, values_only=True):
                                        if not row[0]:
                                            continue  # Skip empty rows

                                        deleted_at = row[deleted_at_col] if deleted_at_col < len(row) else None

                                        # Check if this markup is deleted (has a deleted_at timestamp)
                                        if deleted_at and str(deleted_at).strip() and str(deleted_at) != 'None':
                                            filename = row[filename_col] if filename_col < len(row) else 'unknown'
                                            report_content = row[report_content_col] if report_content_col < len(row) else ''

                                            deleted_markups.append({
                                                'filename': filename,
                                                'description': report_content,
                                                'deleted_at': str(deleted_at)
                                            })

                                wb.close()
                            except ImportError:
                                pass  # openpyxl not available, skip XLSX files
                            except Exception as e:
                                print(f"Warning: Could not read {markup_log_xlsx}: {e}")

                        # Try CSV (for non-XLSX workflows)
                        elif markup_log_csv.exists():
                            try:
                                with open(markup_log_csv, 'r') as log_f:
                                    log_reader = csv.DictReader(log_f)
                                    for log_row in log_reader:
                                        deleted_at = log_row.get('deleted_at', '').strip()

                                        # Check if this markup is deleted
                                        if deleted_at and deleted_at != 'None':
                                            deleted_markups.append({
                                                'filename': log_row.get('filename', 'unknown'),
                                                'description': log_row.get('content', '')[:100],  # First 100 chars
                                                'deleted_at': deleted_at
                                            })
                            except Exception as e:
                                print(f"Warning: Could not read {markup_log_csv}: {e}")

                        # If we found deleted markups, add to global log
                        if deleted_markups:
                            DELETED_MARKUPS_LOG.append({
                                'workflow': workflow['name'],
                                'report_number': report_num,
                                'session_path': str(relative_path),
                                'deleted_markups': deleted_markups,
                                'total_deleted': len(deleted_markups)
                            })

                # This test doesn't fail - it just collects information
                self.assertTrue(True, "Deleted markups logged successfully")


class TestAnnotationWorkflowConsistency(unittest.TestCase):
    """Test consistency between different workflow versions."""

    def setUp(self):
        """Set up test fixtures."""
        global CUSTOM_DIRECTORY, CUSTOM_WORKFLOW

        # Don't run consistency tests if testing a single workflow
        if CUSTOM_WORKFLOW:
            self.skipTest("Consistency tests skipped when testing single workflow")

        # Determine the base directory
        if CUSTOM_DIRECTORY:
            base_dir = Path(CUSTOM_DIRECTORY)
        else:
            base_dir = Path(__file__).parent

        self.mac_xlsx = base_dir / 'annotation_mac_xlsx'
        self.windows_xlsx = base_dir / 'annotation_windows_xlsx'

    def test_mac_and_windows_xlsx_have_same_patients(self):
        """Test that Mac and Windows XLSX versions have identical patient structures."""
        if not (self.mac_xlsx.exists() and self.windows_xlsx.exists()):
            self.skipTest("Both Mac and Windows XLSX workflows not found")

        mac_log = self.mac_xlsx / 'log.csv'
        windows_log = self.windows_xlsx / 'log.csv'

        if not (mac_log.exists() and windows_log.exists()):
            self.skipTest("log.csv not found in both workflows")

        # Read paths from both logs
        def read_log_paths(log_path):
            paths = []
            with open(log_path, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    path = row.get('Path', '').strip()
                    if path.startswith('./'):
                        path = path[2:]
                    paths.append(path)
            return sorted(paths)

        mac_paths = read_log_paths(mac_log)
        windows_paths = read_log_paths(windows_log)

        self.assertEqual(
            mac_paths, windows_paths,
            "Mac and Windows XLSX workflows should have identical patient structures"
        )


def write_deleted_markups_log(output_file='deleted_markups_log.txt'):
    """Write the deleted markups log to a file."""
    global DELETED_MARKUPS_LOG

    if not DELETED_MARKUPS_LOG:
        return None  # No deleted markups found

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    with open(output_file, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write("DELETED MARKUPS REPORT\n")
        f.write(f"Generated: {timestamp}\n")
        f.write("=" * 80 + "\n\n")

        total_reports_with_deletions = len(DELETED_MARKUPS_LOG)
        total_deleted_markups = sum(item['total_deleted'] for item in DELETED_MARKUPS_LOG)

        f.write(f"Summary:\n")
        f.write(f"  - Reports with deleted markups: {total_reports_with_deletions}\n")
        f.write(f"  - Total deleted markups: {total_deleted_markups}\n\n")

        f.write("=" * 80 + "\n\n")

        for entry in DELETED_MARKUPS_LOG:
            f.write(f"Workflow: {entry['workflow']}\n")
            f.write(f"Report Number: {entry['report_number']}\n")
            f.write(f"Path: {entry['session_path']}\n")
            f.write(f"Total Deleted Markups: {entry['total_deleted']}\n")
            f.write("-" * 80 + "\n")

            for i, markup in enumerate(entry['deleted_markups'], 1):
                f.write(f"  [{i}] {markup['filename']}\n")
                if markup['description']:
                    f.write(f"      Description: {markup['description']}\n")
                f.write(f"      Deleted at: {markup['deleted_at']}\n")
                f.write("\n")

            f.write("=" * 80 + "\n\n")

    return output_file


def write_deleted_markups_json(output_file='deleted_markups_log.json'):
    """Write the deleted markups log to a JSON file."""
    global DELETED_MARKUPS_LOG

    if not DELETED_MARKUPS_LOG:
        return None

    output = {
        'generated_at': datetime.now().isoformat(),
        'summary': {
            'reports_with_deletions': len(DELETED_MARKUPS_LOG),
            'total_deleted_markups': sum(item['total_deleted'] for item in DELETED_MARKUPS_LOG)
        },
        'reports': DELETED_MARKUPS_LOG
    }

    with open(output_file, 'w') as f:
        json.dump(output, f, indent=2)

    return output_file


def run_tests(verbosity=2):
    """Run all tests with specified verbosity."""
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestAnnotationWorkflowIntegrity))
    suite.addTests(loader.loadTestsFromTestCase(TestAnnotationWorkflowConsistency))

    # Run tests
    runner = unittest.TextTestRunner(verbosity=verbosity)
    result = runner.run(suite)

    return result


def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='Test annotation workflow data integrity',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Test workflows in current directory
  python test_annotation_workflow.py

  # Test specific directory containing multiple workflows
  python test_annotation_workflow.py --directory /path/to/workflows

  # Test a single workflow directory
  python test_annotation_workflow.py --workflow /path/to/annotation_mac_xlsx

  # Verbose output
  python test_annotation_workflow.py -v

  # Quiet output
  python test_annotation_workflow.py -q
        """
    )

    parser.add_argument(
        '--directory', '-d',
        type=str,
        help='Directory containing annotation workflow folders (e.g., annotation_mac_xlsx, annotation_windows_xlsx)'
    )

    parser.add_argument(
        '--workflow', '-w',
        type=str,
        help='Path to a single workflow directory to test'
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Verbose output (show all test details)'
    )

    parser.add_argument(
        '-q', '--quiet',
        action='store_true',
        help='Quiet output (minimal output)'
    )

    return parser.parse_args()


if __name__ == '__main__':
    # Parse command-line arguments
    args = parse_args()

    # Set global variables
    CUSTOM_DIRECTORY = args.directory
    CUSTOM_WORKFLOW = args.workflow

    # Validate arguments
    if args.directory and args.workflow:
        print("ERROR: Cannot specify both --directory and --workflow")
        sys.exit(1)

    if args.directory:
        dir_path = Path(args.directory)
        if not dir_path.exists():
            print(f"ERROR: Directory does not exist: {args.directory}")
            sys.exit(1)
        if not dir_path.is_dir():
            print(f"ERROR: Path is not a directory: {args.directory}")
            sys.exit(1)
        print(f"Testing workflows in: {dir_path.absolute()}")

    if args.workflow:
        workflow_path = Path(args.workflow)
        if not workflow_path.exists():
            print(f"ERROR: Workflow directory does not exist: {args.workflow}")
            sys.exit(1)
        if not workflow_path.is_dir():
            print(f"ERROR: Path is not a directory: {args.workflow}")
            sys.exit(1)
        print(f"Testing single workflow: {workflow_path.absolute()}")

    # Determine verbosity
    if args.quiet:
        verbosity = 0
    elif args.verbose:
        verbosity = 2
    else:
        verbosity = 1

    # Run tests
    result = run_tests(verbosity=verbosity)

    # Write deleted markups logs if any were found
    if DELETED_MARKUPS_LOG:
        print("\n" + "=" * 80)
        print(f"Found {len(DELETED_MARKUPS_LOG)} report(s) with deleted markups")
        print("=" * 80)

        # Write text log
        txt_file = write_deleted_markups_log()
        if txt_file:
            print(f"✓ Detailed report written to: {txt_file}")

        # Write JSON log
        json_file = write_deleted_markups_json()
        if json_file:
            print(f"✓ JSON report written to: {json_file}")

        print("=" * 80 + "\n")
    else:
        print("\n✓ No deleted markups found in any reports\n")

    # Exit with appropriate code
    sys.exit(0 if result.wasSuccessful() else 1)
