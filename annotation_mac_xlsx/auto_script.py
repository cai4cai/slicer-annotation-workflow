# --- Import required modules ---
import os
import argparse
import slicer
import qt
from qt import QTimer
import datetime
import json
import glob
import errno
import re

try:
    import pandas as pd
except ImportError:
    slicer.util.pip_install("pandas")
    import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    slicer.util.pip_install("openpyxl")
    from openpyxl import Workbook, load_workbook


# --- Configure application settings ---
settings = slicer.app.settings()
settings.setValue("SubjectHierarchy/ResetFieldOfViewOnShowVolume", False)
settings.setValue("SubjectHierarchy/ResetViewOrientationOnShowVolume", False)


# --- Parse command line arguments from Slicer ---
parser = argparse.ArgumentParser()
parser.add_argument('--source_folder', required=True)  # Source directory for data
parser.add_argument('--nifti_files', nargs='*', default=[])  # List of NIfTI file paths
parser.add_argument('--markup_files', nargs='*', default=[])  # List of markup JSON files
parser.add_argument('--report_number', required=True)  # Report number for log identification
parser.add_argument('--log_csv', required=True)  # Path to main log CSV file
args = parser.parse_args()

# --- Define paths and initialise variables ---
sourceFolder = os.path.abspath(args.source_folder)
nifti_files = [os.path.normpath(p) for p in args.nifti_files]
markup_files = [os.path.normpath(p) for p in args.markup_files]
markup_log_file = os.path.join(sourceFolder, "markup_log.xlsx")
loaded_markup_names = []  # Track loaded markup names

# --- Set default style as Light ---
slicer.app.setStyle("Light Slicer")

# --- Function to load NIfTI and markup files into Slicer ---
def loadEverything():
    if nifti_files:
        for nifti_path in nifti_files:
            print(f"Loading volume: {nifti_path}")
            slicer.util.loadVolume(nifti_path)
    else:
        print("No NIfTI files provided.")

    if markup_files:
        print("Loading markup files...")
        for markup_path in markup_files:
            if markup_path and os.path.isfile(markup_path):
                try:
                    slicer.util.loadMarkups(markup_path)
                    loaded_markup_names.append(os.path.basename(markup_path).replace(".json", ""))
                    print(f"Loaded markup: {markup_path}")
                except Exception as e:
                    print(f"Failed to load markup {markup_path}: {e}")
            else:
                print(f"Markup file not found: {markup_path}")
    else:
        print("No markup files provided.")

    slicer.util.selectModule('Data')
    print("Ready for annotation.")

    setAnatomicalSliceViews()

# --- Set 3D slice visibility on the fourth viewer ---
# Get all slice composite nodes (which control slice visibility in 3D)
scene = slicer.mrmlScene
sliceNodes = scene.GetNodesByClass("vtkMRMLSliceNode")

# Toggle visibility for each slice in 3D view
sliceNodes.InitTraversal()
for i in range(sliceNodes.GetNumberOfItems()):
    sliceNode = sliceNodes.GetItemAsObject(i)
    sliceNode.SetSliceVisible(True)
    # currentVisibility = sliceNode.GetSliceVisible()
    # sliceNode.SetSliceVisible(not currentVisibility)  # Toggle: 1 → 0, 0 → 1

# --- Global references for UI components ---
moduleDropdown = None
customToolBar = None

# --- Add a module selection dropdown to the toolbar ---
def addAllowedModuleDropdown(moduleToolBar, allowedModules=None):
    global moduleDropdown
    if moduleToolBar is None:
        return
    if allowedModules is None:
        allowedModules = ['Data', 'Markups', 'Volumes']

    moduleDropdown = qt.QComboBox(moduleToolBar)
    moduleDropdown.addItems(allowedModules)

    # Event handler to switch module when dropdown selection changes
    def onModuleSelected(index):
        moduleName = moduleDropdown.itemText(index)
        slicer.util.selectModule(moduleName)
        print(f"Switched to module: {moduleName}")

    moduleDropdown.currentIndexChanged.connect(onModuleSelected)
    moduleToolBar.addWidget(moduleDropdown)
    slicer.util.selectModule(allowedModules[0])
    print(f"Available modules now: {allowedModules}")
    return moduleDropdown

# --- Initialise custom UI including favourites and dropdown ---
def initialiseCustomUI():
    global customToolBar
    mw = slicer.util.mainWindow()
    if not mw:
        print("Main window not found, cannot initialise UI")
        return

    # Define favourites list
    favourites = ['Data', 'Markups', 'Volumes']

    # Create custom toolbar with dropdown next to it
    customToolBar = qt.QToolBar("CustomModuleToolbar", mw)
    mw.addToolBar(customToolBar)

    # Add dropdown with the favourite modules
    addAllowedModuleDropdown(customToolBar, favourites)

    customToolBar.show()
    print("Custom module dropdown added.")

# --- Clean up custom UI elements on exit ---
def cleanUpCustomUI():
    global moduleDropdown, customToolBar
    if moduleDropdown:
        try:
            moduleDropdown.currentIndexChanged.disconnect()
        except Exception:
            pass
        moduleDropdown.deleteLater()
        moduleDropdown = None
    if customToolBar:
        mw = slicer.util.mainWindow()
        if mw:
            mw.removeToolBar(customToolBar)
        customToolBar.deleteLater()
        customToolBar = None
    print("Custom UI cleaned up.")

# --- Extract markup content as JSON ---
def extract_markup_content(node):
    try:
        markups_data = {}
        n_points = node.GetNumberOfControlPoints()
        points = []
        for i in range(n_points):
            ras_position = [0.0, 0.0, 0.0]
            node.GetNthControlPointPosition(i, ras_position)
            lps_position = [-ras_position[0], -ras_position[1], ras_position[2]]
            label = node.GetNthControlPointLabel(i)
            points.append({'label': label, 'position': lps_position})

        markups_data['points'] = points
        markups_data['name'] = node.GetName()
        markups_data['id'] = node.GetID()
        markups_data['number_of_points'] = n_points

        # Add coordinate system (Slicer uses LPS for export typically)
        markups_data['coordinateSystem'] = "LPS"

        # Add associated parent transform, if any
        parent_transform = node.GetParentTransformNode()
        markups_data['parentTransformID'] = parent_transform.GetID() if parent_transform else None

        # If it's an ROI node, extract its size
        if node.IsA("vtkMRMLMarkupsROINode"):
            try:
                size = [0.0, 0.0, 0.0]
                node.GetRadiusXYZ(size)#.GetSize(size)
                size = [s * 2 for s in size] # GetRadiusXYZ() ×2 gives the same size as when ROI saved normally
                markups_data['size'] = size
            except AttributeError:
                print(f"ROI node {node.GetName()} has no GetSize() method.")
                markups_data['size'] = None

        return json.dumps(markups_data)

    except Exception as extract_error:
        print(f"Failed to extract markup content for {node.GetName()}: {extract_error}")
        return "Failed to extract content"

# --- Save markups, update logs, and clean UI on application exit ---
def onAppExit(caller=None, event=None):
    print("Slicer is closing. Cleaning up custom UI...")
    cleanUpCustomUI()

    print("Slicer is closing. Saving all markups...")

    final_markup_nodes = slicer.util.getNodesByClass('vtkMRMLMarkupsNode')

    # Load existing markup log
    markup_log = {}
    markup_error = {}

    if os.path.exists(markup_log_file):
        wb = load_workbook(markup_log_file, read_only=True)
        ws = wb.active  # First sheet by default

        # Get header → column index mapping
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=0)}

        # Go through rows after header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue  # Skip empty first column

            markup_log[row[0]] = {
                'report_content': row[headers['report_content']],
                'original_filename': row[headers['original_filename']],
                'created_at': row[headers['created_at']],
                'deleted_at': row[headers['deleted_at']]
            }

    # Get last number if any
    max_number = 0
    for markup in markup_log.keys():
        try:
            number_part = int(markup.replace(".json", "").split("_")[-1])
            if number_part > max_number:
                max_number = number_part
        except ValueError:
            continue

    now_str = datetime.datetime.now().isoformat()
    current_filenames = set()

    for markupNode in final_markup_nodes:
        node_name = markupNode.GetName().strip()

        # Check for existing
        existing_file = None
        for fname in markup_log.keys():
            if fname.replace(".json", "") == node_name: 
                existing_file = fname
                break

        if existing_file:
            safe_name = existing_file
        else:
            max_number += 1
            class_name = markupNode.GetClassName()
            if class_name == "vtkMRMLMarkupsFiducialNode":
                markup_type = "point"
            elif class_name == "vtkMRMLMarkupsLineNode":
                markup_type = "line"
            elif class_name == "vtkMRMLMarkupsROINode":
                markup_type = "roi"
            safe_name = f"{markup_type}_{max_number}.json"

        output_path = os.path.join(sourceFolder, safe_name)
        current_filenames.add(safe_name)

        try:
            if existing_file:
                # Save markup
                slicer.util.saveNode(markupNode, output_path)
                print(f"Saved markup: {output_path}")
            # If not previously on markup_log then add
            else:
                markupNode.SetDescription(node_name)
                slicer.util.saveNode(markupNode, output_path)
                markup_log[safe_name] = {
                        'report_content': node_name,
                        'original_filename': None,
                        'created_at': now_str,
                        'deleted_at': None
                }
                print(f"Saved markup: {output_path}")

        except OSError as e:
            print(f"Error saving markup {node_name}: {e}")
            markup_error[safe_name] = {
                        'report_content': node_name,
                        'original_filename': None,
                        'created_at': now_str,
                        'deleted_at': None,
                        'json_content': extract_markup_content(markupNode)
                    }
            
    # Detect deleted markups from log
    saved_filenames = set(markup_log.keys())
    deleted_files = saved_filenames - current_filenames

    for deleted_file in deleted_files:
        file_path = os.path.join(sourceFolder, deleted_file)
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Deleted markup file: {deleted_file}")
            if deleted_file in markup_log:
                if not markup_log[deleted_file].get('deleted_at'):
                    markup_log[deleted_file]['deleted_at'] = now_str
        except Exception as e:
            print(f"Failed to delete {deleted_file}: {e}")

    # Create a new workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Markup Log"

    # Write header
    ws.append(["filename", "report_content", "original_filename", "created_at", "deleted_at"])

    # Write data rows
    for filename, details in markup_log.items():
        ws.append([
            filename,
            details.get('report_content', ''),
            details.get('original_filename', ''),
            details.get('created_at', ''),
            details.get('deleted_at', '')
        ])

        # Save to file
        wb.save(markup_log_file)
    
    # Create markup error log with content
    if markup_error:
        # Sanitize now_str for filename (replace colon with dash)
        safe_timestamp = re.sub(r'[:]', '-', now_str)
        error_filename = f"{safe_timestamp}_error.xlsx"
        error_filepath = os.path.join(sourceFolder, error_filename)

        wb_error = Workbook()
        ws_error = wb_error.active
        ws_error.title = "Markup Errors"

        # Write header including json_content
        ws_error.append(["filename", "report_content", "original_filename", "created_at", "deleted_at", "json_content"])

        for filename, details in markup_error.items():
            ws_error.append([
                filename,
                details.get('report_content', ''),
                details.get('original_filename', ''),
                details.get('created_at', ''),
                details.get('deleted_at', ''),
                details.get('json_content', '')
            ])

        wb_error.save(error_filepath)
        print(f"Saved markup errors to: {error_filepath}")


# --- UI customisation routines ---
def hideAllGUIComponents():

    def mainWindow():
        return slicer.util.mainWindow()

    def findChild(parent, name):
        return parent.findChild(qt.QWidget, name)

    mw = mainWindow()
    if mw is None:
        print("No main window found.")
        return

    # Hide status bar
    if mw.statusBar():
        mw.statusBar().setVisible(False)

    # Hide Python console
    console = mw.pythonConsole().parent()
    if console:
        console.setVisible(False)

    # Hide Error log
    mw.errorLogDockWidget().setVisible(False)

    # Hide module panel title bar
    modulePanelDockWidgets = mw.findChildren(qt.QDockWidget, "PanelDockWidget")
    for dock in modulePanelDockWidgets:
        dock.setTitleBarWidget(qt.QWidget(dock))

    # Hide Help section
    modulePanel = findChild(mw, "ModulePanel")
    if modulePanel:
        modulePanel.helpAndAcknowledgmentVisible = False

    # Hide Data probe
    dataProbeWidget = findChild(mw, "DataProbeCollapsibleWidget")
    if dataProbeWidget:
        dataProbeWidget.setVisible(False)

# --- Ensure that anatomical coordinates are used ---
def setAnatomicalSliceViews():
    layoutManager = slicer.app.layoutManager()
    sliceViewOrientations = {
        "Red": "Axial",
        "Yellow": "Sagittal",
        "Green": "Coronal"
    }

    for sliceName, orientation in sliceViewOrientations.items():
        try:
            sliceWidget = layoutManager.sliceWidget(sliceName)
            if not sliceWidget:
                print(f"[Slice Orientation] Slice widget '{sliceName}' not found.")
                continue
            sliceNode = sliceWidget.sliceLogic().GetSliceNode()
            sliceNode.SetOrientation(orientation)
            sliceNode.SetSliceVisible(True)
            print(f"[Slice Orientation] Set {sliceName} to {orientation}")
        except Exception as e:
            print(f"[Slice Orientation] Failed to set {sliceName} to {orientation}: {e}")

# --- Custom widget to display report on the right ---
def show_report_dock(report_file_path, report_number):
    mw = slicer.util.mainWindow()

    existingDock = mw.findChild(qt.QDockWidget, "ReportDock")
    if existingDock:
        existingDock.close()
        mw.removeDockWidget(existingDock)

    dockWidget = qt.QDockWidget()
    dockWidget.objectName = "ReportDock"
    dockWidget.windowTitle = f"Report: {report_number}"

    dockContents = qt.QWidget()
    dockLayout = qt.QVBoxLayout(dockContents)

    label = qt.QLabel(f"Report for {report_number}")
    label.setStyleSheet("font-weight: bold; font-size: 16px; margin-bottom: 10px;")
    dockLayout.addWidget(label)

    textBrowser = qt.QTextBrowser()
    textBrowser.setReadOnly(True)
    textBrowser.setStyleSheet("background-color: #f0f0f0; font-size:14px")

    if os.path.exists(report_file_path):
        with open(report_file_path, 'r') as f:
            textBrowser.setPlainText(f.read())
    else:
        textBrowser.setPlainText(f"Report file not found: {report_file_path}")

    dockLayout.addWidget(textBrowser)

    dockWidget.setWidget(dockContents)
    mw.addDockWidget(qt.Qt.RightDockWidgetArea, dockWidget)
    dockWidget.show()

# --- Custom widget to display the correlation between markups and report content ---
def show_excel_preview_dock(excel_path):
    mw = slicer.util.mainWindow()

    existingDock = mw.findChild(qt.QDockWidget, "ExcelPreviewDock")
    if existingDock:
        existingDock.close()
        mw.removeDockWidget(existingDock)

    dockWidget = qt.QDockWidget()
    dockWidget.objectName = "ExcelPreviewDock"
    dockWidget.windowTitle = "Markup Log Overview"

    dockContents = qt.QWidget()
    dockLayout = qt.QVBoxLayout(dockContents)

    tableWidget = qt.QTableWidget()
    dockLayout.addWidget(tableWidget)

    if os.path.exists(excel_path):
        try:
            df = pd.read_excel(excel_path, engine='openpyxl')
            df_preview = df.iloc[:, [0,1,3,4]]  # First two columns

            tableWidget.setColumnCount(len(df_preview.columns))
            tableWidget.setRowCount(len(df_preview))

            # Set header labels
            tableWidget.setHorizontalHeaderLabels(df_preview.columns.tolist())

            # Fill table cells
            for row_idx in range(len(df_preview)):
                for col_idx in range(len(df_preview.columns)):
                    cell_value = str(df_preview.iat[row_idx, col_idx])
                    item = qt.QTableWidgetItem(cell_value)
                    tableWidget.setItem(row_idx, col_idx, item)

            tableWidget.resizeColumnsToContents()

        except Exception as e:
            error_label = qt.QLabel(f"Failed to load Excel file: {e}")
            dockLayout.addWidget(error_label)
    else:
        error_label = qt.QLabel("Excel file not found at the moment.")
        dockLayout.addWidget(error_label)

    dockWidget.setWidget(dockContents)
    mw.addDockWidget(qt.Qt.RightDockWidgetArea, dockWidget)
    dockWidget.show()


# report_file_path = os.path.join(args.source_folder, f"report_{args.report_number}.txt")
# Find all files matching report_*.txt in the source folder
report_files = glob.glob(os.path.join(args.source_folder, "report_*.txt"))

# If you expect only one match and want to get it:
if report_files:
    report_file_path = report_files[0]  # or handle multiple files if needed
else:
    report_file_path = None  # or raise an error / handle missing file

# --- Bind exit event and initialise app ---
slicer.app.connect("aboutToQuit()", onAppExit)
QTimer.singleShot(50, hideAllGUIComponents)
QTimer.singleShot(100, loadEverything)
QTimer.singleShot(150, initialiseCustomUI)
QTimer.singleShot(200, lambda: show_report_dock(report_file_path, args.report_number))
QTimer.singleShot(250, lambda: show_excel_preview_dock(markup_log_file))

